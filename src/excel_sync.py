"""
Excel synchronization module for teaching software manager.
Reads and writes data to/from Excel files (openpyxl).
Supports bidirectional sync between YAML and Excel.
"""

from pathlib import Path
from typing import Dict, List, Any, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import yaml

from config_loader import ConfigLoader


class ExcelSyncManager:
    """Manage synchronization between YAML config and Excel file."""

    def __init__(self, config_loader: ConfigLoader, excel_path: str = None):
        """
        Initialize Excel sync manager.
        
        Args:
            config_loader: ConfigLoader instance
            excel_path: Path to Excel file (default: config/teaching_software.xlsx)
        """
        self.config_loader = config_loader
        
        if excel_path is None:
            base_dir = Path(__file__).parent.parent
            excel_path = base_dir / "config" / "teaching_software.xlsx"
        
        self.excel_path = Path(excel_path)

    def _ensure_excel_exists(self) -> bool:
        """Create Excel file if it doesn't exist."""
        if self.excel_path.exists():
            return True
        
        # Create new workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create sheets
        self._create_instructors_sheet(wb)
        self._create_modules_sheet(wb)
        self._create_software_sheet(wb)
        self._create_software_by_os_sheet(wb)
        # Optional ChangeLog sheet for audit trail
        try:
            self._create_changelog_sheet(wb)
        except Exception:
            pass
        
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.excel_path)
        return True

    def _create_instructors_sheet(self, wb: openpyxl.Workbook) -> None:
        """Create Instructors sheet template."""
        ws = wb.create_sheet("Instructors")
        
        headers = ["ID", "Name", "Email", "Department", "Modules", "Last Review"]
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 20

    def _create_modules_sheet(self, wb: openpyxl.Workbook) -> None:
        """Create Modules sheet template."""
        ws = wb.create_sheet("Modules")
        
        headers = ["ID", "Code", "Name", "Description", "Year", "Semester", "OS Required", "Instructor"]
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 20

    def _create_software_sheet(self, wb: openpyxl.Workbook) -> None:
        """Create Software sheet template."""
        ws = wb.create_sheet("Software")
        
        headers = ["Module ID", "Software Name", "Version", "Purpose", "Category", "Critical", "OS Supported", "Notes", "Last Verified", "Verified By"]
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 35
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 20

    def _create_changelog_sheet(self, wb: openpyxl.Workbook) -> None:
        """Create ChangeLog sheet template for audit trail."""
        ws = wb.create_sheet("ChangeLog")
        headers = ["Timestamp", "Module ID", "Software Name", "Action", "Actor", "Field", "Old Value", "New Value"]
        ws.append(headers)
        header_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 22
        ws.column_dimensions['H'].width = 22

    def _create_software_by_os_sheet(self, wb: openpyxl.Workbook) -> None:
        """Create Software by OS sheet template."""
        ws = wb.create_sheet("Software by OS")
        
        headers = ["OS", "Software Name", "Version", "Category", "Module ID", "Purpose", "Critical", "Notes"]
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 30

    def export_to_excel(self) -> Tuple[bool, str]:
        """
        Export YAML config to Excel file.
        
        Returns:
            Tuple of (success, message)
        """
        try:
            self._ensure_excel_exists()
            wb = openpyxl.load_workbook(self.excel_path)
            
            # Export instructors
            instructors = self.config_loader.get_instructors()
            ws_inst = wb["Instructors"]
            # Clear existing data, but only if there are data rows
            if ws_inst.max_row > 1:
                ws_inst.delete_rows(2, ws_inst.max_row - 1)
            
            for inst_id, inst_data in instructors.items():
                modules_str = ", ".join(inst_data.get('modules', []))
                ws_inst.append([
                    inst_id,
                    inst_data.get('name', ''),
                    inst_data.get('email', ''),
                    inst_data.get('department', ''),
                    modules_str,
                    inst_data.get('last_review', '')
                ])
            
            # Export modules
            modules = self.config_loader.get_modules()
            ws_mod = wb["Modules"]
            if ws_mod.max_row > 1:
                ws_mod.delete_rows(2, ws_mod.max_row - 1)
            
            for mod_id, mod_data in modules.items():
                instructor = mod_data.get('instructor_id', '')
                os_required = []
                for os_entry in mod_data.get('os_required', []):
                    if isinstance(os_entry, dict):
                        name = os_entry.get('name') or ''
                        note = os_entry.get('note') or os_entry.get('description') or ''
                        os_required.append(f"{name} ({note})".strip()) if note else os_required.append(name)
                    else:
                        os_required.append(str(os_entry))
                os_required_str = ", ".join([o for o in os_required if o])
                ws_mod.append([
                    mod_id,
                    mod_data.get('code', ''),
                    mod_data.get('name', ''),
                    mod_data.get('description', ''),
                    mod_data.get('year', ''),
                    mod_data.get('semester', ''),
                    os_required_str,
                    instructor
                ])
            
            # Export software
            ws_soft = wb["Software"]
            if ws_soft.max_row > 1:
                ws_soft.delete_rows(2, ws_soft.max_row - 1)
            
            for mod_id, mod_data in modules.items():
                module_os_names = []
                for os_entry in mod_data.get('os_required', []):
                    if isinstance(os_entry, dict):
                        if os_entry.get('name'):
                            module_os_names.append(os_entry.get('name'))
                    else:
                        module_os_names.append(str(os_entry))
                for software in mod_data.get('software', []):
                    os_list = software.get('os_supported', [])
                    if not os_list and module_os_names:
                        os_list = module_os_names
                    os_supported = ', '.join(os_list)
                    ws_soft.append([
                        mod_id,
                        software.get('name', ''),
                        software.get('version', ''),
                        software.get('purpose', ''),
                        software.get('category', ''),
                        'Yes' if software.get('critical', False) else 'No',
                        os_supported,
                        software.get('notes', ''),
                        software.get('last_verified', ''),
                        software.get('verified_by', '')
                    ])
            
            # Export software by OS
            ws_os = wb["Software by OS"] if "Software by OS" in wb.sheetnames else None
            if ws_os is None:
                self._create_software_by_os_sheet(wb)
                ws_os = wb["Software by OS"]
            if ws_os.max_row > 1:
                ws_os.delete_rows(2, ws_os.max_row - 1)
            
            # Build a mapping of OS to software
            os_to_software = {}
            for mod_id, mod_data in modules.items():
                module_os_names = []
                for os_entry in mod_data.get('os_required', []):
                    if isinstance(os_entry, dict):
                        if os_entry.get('name'):
                            module_os_names.append(os_entry.get('name'))
                    else:
                        module_os_names.append(str(os_entry))
                for software in mod_data.get('software', []):
                    os_list = software.get('os_supported', [])
                    if not os_list and module_os_names:
                        os_list = module_os_names
                    for os_name in os_list:
                        if os_name not in os_to_software:
                            os_to_software[os_name] = []
                        os_to_software[os_name].append({
                            'name': software.get('name', ''),
                            'version': software.get('version', ''),
                            'category': software.get('category', ''),
                            'module_id': mod_id,
                            'purpose': software.get('purpose', ''),
                            'critical': software.get('critical', False),
                            'notes': software.get('notes', '')
                        })
            
            # Sort OS names and export
            for os_name in sorted(os_to_software.keys()):
                for soft_info in os_to_software[os_name]:
                    ws_os.append([
                        os_name,
                        soft_info['name'],
                        soft_info['version'],
                        soft_info['category'],
                        soft_info['module_id'],
                        soft_info['purpose'],
                        'Yes' if soft_info['critical'] else 'No',
                        soft_info['notes']
                    ])
            
            # Export audit log to ChangeLog
            try:
                ws_log = wb["ChangeLog"] if "ChangeLog" in wb.sheetnames else None
                if ws_log is None:
                    self._create_changelog_sheet(wb)
                    ws_log = wb["ChangeLog"]
                if ws_log.max_row > 1:
                    ws_log.delete_rows(2, ws_log.max_row - 1)
                audit_log = self.config_loader.config.get('audit_log', [])
                for entry in audit_log:
                    ts = entry.get('timestamp', '')
                    mod_id = entry.get('module_id', '')
                    soft = entry.get('software_name', '')
                    action = entry.get('action', '')
                    actor = entry.get('actor', '')
                    if action == 'updated':
                        for ch in entry.get('changes', []):
                            ws_log.append([ts, mod_id, soft, action, actor, ch.get('field', ''), ch.get('old', ''), ch.get('new', '')])
                    else:
                        ws_log.append([ts, mod_id, soft, action, actor, '*', '', ''])
            except Exception:
                # Non-blocking: if log export fails, continue
                pass

            try:
                wb.save(self.excel_path)
            except PermissionError:
                return False, f"Error: Cannot write to {self.excel_path.name}. Is the file open in Excel? Please close it and try again."

            return True, f"✓ Exported to {self.excel_path.name}"
        
        except Exception as e:
            return False, f"Error exporting to Excel: {str(e)}"

    def import_from_excel(self) -> Tuple[bool, str]:
        """
        Import data from Excel file to YAML config.
        
        Returns:
            Tuple of (success, message)
        """
        try:
            if not self.excel_path.exists():
                return False, f"Excel file not found: {self.excel_path}"
            
            wb = openpyxl.load_workbook(self.excel_path)
            config = self.config_loader.config
            
            # Import instructors
            instructors = {}
            ws_inst = wb["Instructors"] if "Instructors" in wb.sheetnames else None
            if ws_inst:
                for row in ws_inst.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # ID exists
                        modules_str = row[4] if row[4] else ""
                        modules = [m.strip() for m in modules_str.split(",") if m.strip()]
                        
                        instructors[row[0]] = {
                            'name': row[1] or '',
                            'email': row[2] or '',
                            'department': row[3] or '',
                            'modules': modules,
                            'last_review': row[5] or ''
                        }
            
            # Import modules
            modules = {}
            ws_mod = wb["Modules"] if "Modules" in wb.sheetnames else None
            if ws_mod:
                for row in ws_mod.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # ID exists
                        os_required_list = []
                        os_raw = row[6] if len(row) > 6 else ''
                        if os_raw:
                            for os_item in os_raw.split(','):
                                name = os_item.strip()
                                if name:
                                    os_required_list.append({'name': name})
                        instructor_val = row[7] if len(row) > 7 else ''
                        modules[row[0]] = {
                            'code': row[1] or '',
                            'name': row[2] or '',
                            'description': row[3] or '',
                            'year': int(row[4]) if row[4] else 1,
                            'semester': int(row[5]) if row[5] else 1,
                            'os_required': os_required_list,
                            'instructor_id': instructor_val or '',
                            'software': []
                        }
            
            # Import software
            ws_soft = wb["Software"] if "Software" in wb.sheetnames else None
            if ws_soft:
                for row in ws_soft.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:  # Module ID and software name exist
                        mod_id = row[0]
                        if mod_id in modules:
                            # Parse OS supported from comma-separated string
                            os_str = row[6] if row[6] else ""
                            os_supported = [os.strip() for os in os_str.split(",") if os.strip()] if os_str else []
                            if not os_supported:
                                module_os_names = []
                                for os_entry in modules[mod_id].get('os_required', []):
                                    if isinstance(os_entry, dict):
                                        if os_entry.get('name'):
                                            module_os_names.append(os_entry.get('name'))
                                    else:
                                        module_os_names.append(str(os_entry))
                                os_supported = module_os_names
                            
                            software = {
                                'name': row[1],
                                'version': row[2] or '',
                                'purpose': row[3] or '',
                                'category': row[4] or '',
                                'critical': row[5] == 'Yes' if row[5] else False,
                                'os_supported': os_supported,
                                'notes': row[7] or '',
                                'last_verified': row[8] or '',
                                'verified_by': row[9] or ''
                            }
                            modules[mod_id]['software'].append(software)
            
            # Update config
            config['instructors'] = instructors
            config['modules'] = modules
            
            # Save back to YAML
            config_path = Path(__file__).parent.parent / "config" / "teaching_software.yml"
            with open(config_path, 'w', encoding='utf-8') as f:
                yaml.dump(config, f, default_flow_style=False, allow_unicode=True)
            
            # Reload config
            self.config_loader = ConfigLoader()
            
            return True, f"✓ Imported {len(instructors)} instructors, {len(modules)} modules from Excel"
        
        except Exception as e:
            return False, f"Error importing from Excel: {str(e)}"

    def get_sync_status(self) -> Dict[str, Any]:
        """Get current sync status between YAML and Excel."""
        try:
            yaml_count = {
                'instructors': len(self.config_loader.get_instructors()),
                'modules': len(self.config_loader.get_modules())
            }
            
            excel_count = {'instructors': 0, 'modules': 0}
            
            if self.excel_path.exists():
                try:
                    wb = openpyxl.load_workbook(self.excel_path)
                    ws_inst = wb["Instructors"] if "Instructors" in wb.sheetnames else None
                    if ws_inst:
                        excel_count['instructors'] = max(ws_inst.max_row - 1, 0)
                    ws_mod = wb["Modules"] if "Modules" in wb.sheetnames else None
                    if ws_mod:
                        excel_count['modules'] = max(ws_mod.max_row - 1, 0)
                except Exception as wb_err:
                    # Log the error but continue with status
                    print(f"Warning: Could not read Excel file: {wb_err}")
            
            return {
                'yaml': yaml_count,
                'excel': excel_count,
                'excel_exists': self.excel_path.exists(),
                'excel_path': str(self.excel_path)
            }
        
        except Exception as e:
            # Return valid status structure even on error
            return {
                'yaml': {'instructors': 0, 'modules': 0},
                'excel': {'instructors': 0, 'modules': 0},
                'excel_exists': False,
                'excel_path': str(self.excel_path),
                'error': str(e)
            }
